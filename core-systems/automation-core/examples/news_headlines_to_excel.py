#!/usr/bin/env python3
# automation-core/examples/news_headlines_to_excel.py
# -*- coding: utf-8 -*-

"""
News Headlines → Excel/CSV (industrial-grade, no external deps required)
- Fetch RSS/Atom feeds via urllib with retries and timeouts
- Parse XML via xml.etree, normalize fields across formats
- Filter by time window (--since-hours) and keyword query (--query / --any / --all)
- Deduplicate by canonical URL
- Export to Excel via openpyxl if available, otherwise fallback to CSV
- Configurable via CLI and optional JSON/YAML feeds file (YAML requires PyYAML if used)

Usage examples:
  python news_headlines_to_excel.py \
      --feeds https://feeds.bbci.co.uk/news/rss.xml https://www.reutersagency.com/feed/?best-topics=technology
  python news_headlines_to_excel.py --feeds-file feeds.json --since-hours 24 --query "ai,ethereum" --any
  python news_headlines_to_excel.py --out news.xlsx --limit-per-feed 100 --timeout 10 --retries 2
"""

from __future__ import annotations

import argparse
import contextlib
import csv
import dataclasses
import datetime as dt
import html
import io
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from typing import Iterable, List, Optional, Tuple, Dict, Any

# Optional deps
try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    _HAVE_OPENPYXL = True
except Exception:
    _HAVE_OPENPYXL = False

try:
    # YAML is optional for feeds file
    import yaml  # type: ignore
    _HAVE_YAML = True
except Exception:
    _HAVE_YAML = False

# Timezone
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except Exception:
    ZoneInfo = None  # type: ignore


_DEFAULT_UA = "Mozilla/5.0 (compatible; NewsToExcel/1.0; +https://example.org)"
_DATE_RE = re.compile(r"\d{1,2}\s\w{3,}\s\d{4}|\d{4}-\d{2}-\d{2}")
_WHITESPACE_RE = re.compile(r"\s+")
_TAGS_RE = re.compile(r"<[^>]+>")


@dataclasses.dataclass
class FeedItem:
    source: str
    title: str
    url: str
    published: Optional[dt.datetime]
    summary: str


def _now_utc() -> dt.datetime:
    return dt.datetime.now(dt.timezone.utc)


def _to_tz(d: dt.datetime, tz_name: Optional[str]) -> dt.datetime:
    if not d.tzinfo:
        d = d.replace(tzinfo=dt.timezone.utc)
    if tz_name and ZoneInfo:
        try:
            return d.astimezone(ZoneInfo(tz_name))
        except Exception:
            return d.astimezone(dt.timezone.utc)
    return d.astimezone(dt.timezone.utc)


def _clean_text(s: str) -> str:
    s = html.unescape(s or "")
    s = _TAGS_RE.sub(" ", s)
    s = _WHITESPACE_RE.sub(" ", s).strip()
    return s


def _canonical_url(u: str) -> str:
    try:
        parts = urllib.parse.urlsplit(u)
        # drop common tracking params
        q = urllib.parse.parse_qsl(parts.query, keep_blank_values=False)
        q = [(k, v) for (k, v) in q if not k.lower().startswith(("utm_", "gclid", "fbclid"))]
        query = urllib.parse.urlencode(q)
        parts = parts._replace(query=query, fragment="")
        return urllib.parse.urlunsplit(parts)
    except Exception:
        return u.strip()


def _http_get(url: str, timeout: int, retries: int, ua: str) -> bytes:
    last_err: Optional[Exception] = None
    for attempt in range(retries + 1):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": ua, "Accept": "application/rss+xml,application/atom+xml,application/xml;q=0.9,*/*;q=0.8"})
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                if resp.status >= 400:
                    raise urllib.error.HTTPError(url, resp.status, resp.reason, resp.headers, None)
                return resp.read()
        except Exception as e:
            last_err = e
            if attempt < retries:
                time.sleep(min(2 ** attempt, 5))
            else:
                break
    assert last_err is not None
    raise last_err


def _first_text(elem: Optional[ET.Element], paths: Iterable[str]) -> str:
    if elem is None:
        return ""
    for p in paths:
        f = elem.find(p)
        if f is not None and (f.text or "").strip():
            return f.text or ""
    return ""


def _find_any(elem: Optional[ET.Element], tags: Iterable[str]) -> Optional[ET.Element]:
    if elem is None:
        return None
    for t in tags:
        f = elem.find(t)
        if f is not None:
            return f
    return None


def _parse_date(s: str) -> Optional[dt.datetime]:
    s = (s or "").strip()
    if not s:
        return None
    # Try RFC822 / RFC1123 / HTTP date
    for fmt in ("%a, %d %b %Y %H:%M:%S %z", "%d %b %Y %H:%M:%S %z",
                "%a, %d %b %Y %H:%M:%S GMT", "%a, %d %b %Y %H:%M:%S"):
        with contextlib.suppress(Exception):
            d = dt.datetime.strptime(s, fmt)
            if d.tzinfo is None:
                d = d.replace(tzinfo=dt.timezone.utc)
            return d.astimezone(dt.timezone.utc)
    # Try ISO 8601
    with contextlib.suppress(Exception):
        d = dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
        if d.tzinfo is None:
            d = d.replace(tzinfo=dt.timezone.utc)
        return d.astimezone(dt.timezone.utc)
    # Heuristic: extract date part
    m = _DATE_RE.search(s)
    if m:
        with contextlib.suppress(Exception):
            d = dt.datetime.fromisoformat(m.group(0))
            return d.replace(tzinfo=dt.timezone.utc)
    return None


def _parse_rss(content: bytes, default_source: str) -> List[FeedItem]:
    root = ET.fromstring(content)
    ns = {"content": "http://purl.org/rss/1.0/modules/content/"}
    items: List[FeedItem] = []

    channel = _find_any(root, ["channel"])
    source_title = _clean_text(_first_text(channel, ["title"])) or default_source

    for it in root.findall(".//item"):
        title = _clean_text(_first_text(it, ["title"]))
        link = _first_text(it, ["link"])
        pub = _first_text(it, ["pubDate", "published"])
        desc = _first_text(it, ["description", "content:encoded"])
        if not link:
            # Some feeds embed link within <guid isPermaLink="true">
            guid = it.find("guid")
            if guid is not None and (guid.attrib.get("isPermaLink", "").lower() == "true") and (guid.text or "").startswith("http"):
                link = guid.text or ""
        items.append(
            FeedItem(
                source=source_title,
                title=title or "(no title)",
                url=_canonical_url(link),
                published=_parse_date(pub),
                summary=_clean_text(desc),
            )
        )
    return items


def _parse_atom(content: bytes, default_source: str) -> List[FeedItem]:
    root = ET.fromstring(content)
    items: List[FeedItem] = []
    source_title = _clean_text(_first_text(root, ["title"])) or default_source

    for entry in root.findall(".//{*}entry"):
        title = _clean_text(_first_text(entry, ["{*}title"]))
        # link can be in multiple <link> elements; pick rel="alternate" href
        link = ""
        for link_el in entry.findall("{*}link"):
            rel = link_el.attrib.get("rel", "alternate")
            href = link_el.attrib.get("href", "")
            if rel == "alternate" and href:
                link = href
                break
            if href and not link:
                link = href
        pub = _first_text(entry, ["{*}updated", "{*}published"])
        summary = _first_text(entry, ["{*}summary", "{*}content"])
        items.append(
            FeedItem(
                source=source_title,
                title=title or "(no title)",
                url=_canonical_url(link),
                published=_parse_date(pub),
                summary=_clean_text(summary),
            )
        )
    return items


def parse_feed(content: bytes, default_source: str) -> List[FeedItem]:
    # Decide RSS vs Atom by root tag
    root = ET.fromstring(content)
    tag = root.tag.lower()
    if "rss" in tag or tag.endswith("rss") or tag.endswith("rdf"):
        return _parse_rss(content, default_source)
    if "feed" in tag or tag.endswith("feed"):
        return _parse_atom(content, default_source)
    # fallback: try RSS parser
    return _parse_rss(content, default_source)


def load_feeds_from_file(path: str) -> List[str]:
    data: Any
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()
    # Detect JSON vs YAML
    if path.lower().endswith((".yaml", ".yml")):
        if not _HAVE_YAML:
            raise RuntimeError("YAML file provided but PyYAML is not installed.")
        data = yaml.safe_load(text)  # type: ignore
    else:
        data = json.loads(text)
    if isinstance(data, dict):
        feeds = data.get("feeds") or data.get("urls") or data.get("sources")
        if not isinstance(feeds, list):
            raise ValueError("Invalid feeds file structure: expected list under 'feeds'/'urls'/'sources'")
        return [str(x) for x in feeds]
    if isinstance(data, list):
        return [str(x) for x in data]
    raise ValueError("Unsupported feeds file format")


def filter_items(
    items: List[FeedItem],
    since_hours: Optional[int],
    tz_name: Optional[str],
    query: Optional[str],
    match_all: bool,
    limit_per_feed: Optional[int],
) -> List[FeedItem]:
    # Deduplicate by URL; keep the newest published
    by_url: Dict[str, FeedItem] = {}
    cutoff_utc: Optional[dt.datetime] = None
    if since_hours is not None:
        cutoff_utc = _now_utc() - dt.timedelta(hours=since_hours)

    tokens: List[str] = []
    if query:
        tokens = [t.strip() for t in re.split(r"[,\|]", query) if t.strip()]
        tokens = [t.lower() for t in tokens if t]

    # group per source for per-feed limit
    grouped: Dict[str, List[FeedItem]] = {}
    for it in items:
        if cutoff_utc and it.published and it.published < cutoff_utc:
            continue
        if tokens:
            text = f"{it.title} {it.summary}".lower()
            checks = [(tok in text) for tok in tokens]
            if match_all and not all(checks):
                continue
            if not match_all and not any(checks):
                continue

        key = _canonical_url(it.url)
        cur = by_url.get(key)
        if (cur is None) or (it.published and (not cur.published or it.published > cur.published)):
            by_url[key] = it

        grouped.setdefault(it.source, []).append(it)

    # Apply per-feed limit while preserving newest first
    result: List[FeedItem] = []
    for src, lst in grouped.items():
        lst.sort(key=lambda x: (x.published or dt.datetime.min.replace(tzinfo=dt.timezone.utc)), reverse=True)
        if limit_per_feed is not None:
            lst = lst[: max(0, int(limit_per_feed))]
        result.extend(lst)

    # re-apply dedup to final list
    out: List[FeedItem] = []
    seen: set[str] = set()
    for it in sorted(result, key=lambda x: (x.published or dt.datetime.min.replace(tzinfo=dt.timezone.utc)), reverse=True):
        key = _canonical_url(it.url)
        if key in seen:
            continue
        seen.add(key)
        # convert to output timezone for display
        if it.published:
            it = dataclasses.replace(it, published=_to_tz(it.published, tz_name))
        out.append(it)
    return out


def export_excel(path_out: str, rows: List[FeedItem], sheet: str = "News") -> None:
    if not _HAVE_OPENPYXL:
        raise RuntimeError("openpyxl not installed; cannot write .xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet

    headers = ["Published", "Source", "Title", "Summary", "URL"]
    ws.append(headers)
    hdr_font = Font(bold=True)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = hdr_font
        cell.alignment = Alignment(vertical="top")

    for it in rows:
        pub = it.published.isoformat() if it.published else ""
        r = [pub, it.source, it.title, it.summary, it.url]
        ws.append(r)

    # Hyperlink on URL column
    for i in range(2, ws.max_row + 1):
        c = ws.cell(row=i, column=5)
        if c.value:
            c.hyperlink = c.value
            c.style = "Hyperlink"

    # Wrap text and autosize columns
    for col in range(1, 6):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            s = str(v) if v is not None else ""
            max_len = max(max_len, min(len(s), 120))
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(max_len + 2, 120))

    wb.save(path_out)


def export_csv(path_out: str, rows: List[FeedItem]) -> None:
    with open(path_out, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Published", "Source", "Title", "Summary", "URL"])
        for it in rows:
            pub = it.published.isoformat() if it.published else ""
            w.writerow([pub, it.source, it.title, it.summary, it.url])


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Fetch news headlines from RSS/Atom and export to Excel/CSV.")
    p.add_argument("--feeds", nargs="*", help="Feed URLs (space-separated)")
    p.add_argument("--feeds-file", help="JSON or YAML file with list under 'feeds'/'urls'/'sources'")
    p.add_argument("--out", default="news.xlsx", help="Output file (.xlsx or .csv)")
    p.add_argument("--sheet-name", default="News", help="Excel sheet name")
    p.add_argument("--timeout", type=int, default=10, help="HTTP timeout seconds")
    p.add_argument("--retries", type=int, default=2, help="HTTP retries")
    p.add_argument("--user-agent", default=_DEFAULT_UA, help="HTTP User-Agent")
    p.add_argument("--since-hours", type=int, help="Only include items newer than N hours")
    p.add_argument("--tz", default="Europe/Stockholm", help="Output timezone for dates")
    p.add_argument("--query", help="Comma-separated keywords for title/summary match")
    g = p.add_mutually_exclusive_group()
    g.add_argument("--all", action="store_true", help="Match ALL query tokens")
    g.add_argument("--any", action="store_true", help="Match ANY query tokens (default)")
    p.add_argument("--limit-per-feed", type=int, default=100, help="Max items per feed (after filtering)")
    p.add_argument("--verbose", action="store_true", help="Verbose logging")
    args = p.parse_args(argv)

    feeds: List[str] = []
    if args.feeds:
        feeds.extend(args.feeds)
    if args.feeds_file:
        feeds.extend(load_feeds_from_file(args.feeds_file))
    if not feeds:
        print("No feeds provided. Use --feeds or --feeds-file.", file=sys.stderr)
        return 2

    # Fetch + parse
    all_items: List[FeedItem] = []
    for url in feeds:
        try:
            if args.verbose:
                print(f"[FETCH] {url}")
            raw = _http_get(url, timeout=args.timeout, retries=args.retries, ua=args.user_agent)
            items = parse_feed(raw, default_source=urllib.parse.urlsplit(url).netloc or "unknown")
            if args.verbose:
                print(f"  -> {len(items)} items")
            all_items.extend(items)
        except Exception as e:
            print(f"[WARN] Failed {url}: {e}", file=sys.stderr)

    # Filter/dedupe/sort
    match_all = bool(args.all) and not bool(args.any)
    rows = filter_items(
        all_items,
        since_hours=args.since_hours,
        tz_name=args.tz,
        query=args.query,
        match_all=match_all,
        limit_per_feed=args.limit_per_feed,
    )

    # Export
    out_lower = args.out.lower()
    if out_lower.endswith(".xlsx"):
        if not _HAVE_OPENPYXL:
            print("[WARN] openpyxl not installed, falling back to CSV with same basename.", file=sys.stderr)
            out_csv = re.sub(r"\.xlsx$", ".csv", args.out, flags=re.IGNORECASE)
            export_csv(out_csv, rows)
            if args.verbose:
                print(f"[OK] Wrote {len(rows)} rows to {out_csv}")
        else:
            export_excel(args.out, rows, sheet=args.sheet_name)
            if args.verbose:
                print(f"[OK] Wrote {len(rows)} rows to {args.out}")
    else:
        export_csv(args.out, rows)
        if args.verbose:
            print(f"[OK] Wrote {len(rows)} rows to {args.out}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
